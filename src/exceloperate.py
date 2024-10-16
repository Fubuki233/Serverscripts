import os
import openpyxl
import difflib
import tkinter as tk
from tkinter import ttk
import re
stop_row = 1000000
def indexfind():
    global index, cell
    index_unstrip = "NUll"
    for i in cell:
        for j in i:
            if j.value == "编号":
                index_unstrip = i
                break
    if index_unstrip != "Null":
        for k in index_unstrip:
            if k.value != "Null":
                index.append(k)
        for i in range(0, len(index)):
            print(f"index {i} is:{index[i].value}")
    else:
        print("index not found")


# 对比index和target_cell的相似度
def target_index_detect():
    global index, target_cell, target_sheet, target_workbook

    # 初始化保存对比结果的列表
    target_index = []

    target_index_unstrip = None  # 初始化变量

    for i in target_cell:
        for j in i:
            if j.value == "编号":
                target_index_unstrip = i
                break
    if target_index_unstrip:
        for k in target_index_unstrip:
            if k.value:
                target_index.append(k)
    target_sheet = target_workbook[target_workbook.sheetnames[0]]
    target_cell = target_sheet[target_sheet.dimensions]
    target_index = []
    try:
        merge_defendant_columns()
    except Exception as e:
        print("w")
    for i in target_cell:
        for j in i:
            if j.value == "编号":
                target_index_unstrip = i
                break
    if target_index_unstrip:
        for k in target_index_unstrip:
            if k.value:
                target_index.append(k)
    for i in range(len(target_index)):
        print(f"target_index {i} is:{target_index[i].value}")

    # 处理其他索引
    for idx_value in index:
        cleaned_idx_value = idx_value.value.strip()
        for target_value in target_index:
            cleaned_target_value = target_value.value.strip()
            # 计算相似度
            if cleaned_idx_value[:3] == cleaned_target_value[:3]:
                similarity = 1.0
            else:
                # 计算相似度
                similarity = difflib.SequenceMatcher(None, cleaned_idx_value, cleaned_target_value).ratio()
            if similarity > 0.76:
                print(f"相似度: {similarity * 100:.2f}%")
                print(f"源数据: {cleaned_idx_value} | 目标数据: {cleaned_target_value}")
                append_target_data_to_source(idx_value, target_value)





def find_next_available_column(sheet):
    """
    查找表格中下一个空列，用于添加新的合并数据
    """
    max_column = sheet.max_column
    for col in range(1, max_column + 1):
        if sheet.cell(row=1, column=col).value is None:  # 如果该列为空
            return col
    return max_column + 1  # 如果没有空列，则返回下一个新列的索引


def find_last_row(sheet, column):
    """
    查找某一列的最后一行数据，用于追加新数据
    """
    last_row = sheet.max_row
    for row in range(1, sheet.max_row + 1):
        if sheet.cell(row=row, column=column).value is None:
            last_row = row - 1
            break
    return last_row


def merge_defendant_columns():
    global target_sheet
    """
    查找所有列名包含“被告”的列，将这些列的值合并到“被告1”列，并清空其他“被告”列。
    返回合并后的被告数据列表。
    """
    defendant_data = [""] * (target_sheet.max_row + 1)  # 用于存储合并后的被告数据
    defendant_columns = []  # 用于记录所有包含“被告”的列的索引
    save_col = None  # 用于保存“被告1”列的索引

    # 找出所有包含"被告"的列

    for col in range(1, target_sheet.max_column + 1):
        column_header = target_sheet.cell(row=4, column=col).value
        print(column_header)
        if column_header and "被告" in column_header:
            print("Sasdawdawda")
            if column_header == "被告1":
                save_col = col
                print(col)
            else:
                defendant_columns.append(col)
                print("awdawd")
                print(col)
            for row in range(5, target_sheet.max_row + 1):
                if (target_sheet.cell(row=row, column=col).value!=None and target_sheet.cell(row=row, column=col).value!="/"):
                    defendant_name = target_sheet.cell(row=row, column=col).value
                else:
                    defendant_name = ""
                if defendant_data[row]:
                    defendant_data[row] += f"/{defendant_name}"
                else:
                    defendant_data[row] = defendant_name
                defendant_data[row]=defendant_data[row].rstrip("/")



    # 将合并后的数据写入“被告1”列
    if save_col:
        for row, value in enumerate(defendant_data, start=1):
            if row > 5:  # 跳过第一行（列名行）
                try:
                    target_sheet.cell(row=row, column=save_col).value = value
                except Exception as e:
                    print("s")
    # 清空其他“被告”列
    for col in defendant_columns:
        for row in range(2, target_sheet.max_row + 1):
            try:
                target_sheet.cell(row=row, column=col).value = ""
            except Exception as e:
                print("s")
    print(defendant_data)




# 去掉英文括号的函数
def remove_brackets(value):
    if isinstance(value, str):  # 只有字符串才执行替换操作
        return value.replace("(", "").replace(")", "")
    return str(value)  # 非字符串的值直接转换为字符串返回


def is_sequential_and_numeric(value, previous_value):
    # 如果值是数字并且连续（当前值比上一个值大1）
    if value.isdigit() and previous_value.isdigit() and int(value) == int(previous_value) + 1:
        return True
    return False


def append_target_data_to_source(source_cell, target_cell):
    global stop_row
    # 获取源列和目标列的起始坐标
    source_column = source_cell.column
    target_column = target_cell.column

    # 找到源列的最后一个非空行
    source_last_row = find_last_row(sheet, source_column)

    # 初始化前一个编号的值
    previous_value = None
    # 遍历目标列下方的数据，添加到源列
    for row in range(target_cell.row + 1, min(target_sheet.max_row + 1, stop_row)):
        if row > stop_row:
            break
        target_value = target_sheet.cell(row=row, column=target_column).value
        if target_cell.column == 1 and stop_row == 1000000 and not isinstance(target_value, (int, float)) :
            stop_row = row
            print(stop_row)
            continue
        target_value_str = str(target_value).strip()  # 将值转化为字符串并去除空格

        # 如果值不是数字或者不连续，则停止追加数据
        # if previous_value is not None and not is_sequential_and_numeric(target_value_str, previous_value):
        #    print(f"遇到非连续编号或非数字，停止在行: {row}")
        #    break

        # 在源列的最后一行后添加数据
        if target_value is not None:
            sheet.cell(row=source_last_row + 1, column=source_column).value = target_value
            source_last_row += 1
        else:
            sheet.cell(row=source_last_row + 1, column=source_column).value = "/"
            source_last_row += 1


def operate():
    global workbook, target_workbook, sheet, target_sheet, cell, target_cell, index, save_path ,stop_row
    workbook = openpyxl.load_workbook(rf"{module_path.get().strip().replace("\"", "").replace("\"", "")}")
    print(target_path.get().replace("\"", ""))
    target_workbook = openpyxl.load_workbook(rf"{target_path.get().replace("\"", "")}")
    sheet = workbook[workbook.sheetnames[0]]
    target_sheet = target_workbook[target_workbook.sheetnames[0]]
    cell = sheet[sheet.dimensions]
    target_cell = target_sheet[target_sheet.dimensions]
    pattern = r"合同编号：(\w+)"
    match=None
    try:
        match = re.search(pattern, target_sheet.cell(row=2, column=1).value)
        minus = 3
    except Exception as e:
        print("未找到合同编号")
        minus=0
    contract_number=""
    if match:
        contract_number = match.group(1)
        print("提取的合同编号是：", contract_number)
    else:
        print("未找到合同编号")
    index = []
    indexfind()
    target_index_detect()
    file_name_without_extension = os.path.splitext(target_path.get())[0]
    print(file_name_without_extension)
    pattern = r'(\d{4})(\d{2})(\d{2})'
    # 查找所有匹配的日期
    match = re.search(pattern, file_name_without_extension)
    year=""
    season=""
    if match:
        year = match.group(1)
        month = match.group(2)
        print(month)
        day = match.group(3)
        print(f"提取到的月份: {month}")
    else:
        print("未找到符合格式的日期")

    if 1 <= int(month) <= 3:
        season="Q1"
    if 3 < int(month) <= 6:
        season="Q2"
    if 6 < int(month) <= 9:
        season="Q3"
    if 9 < int(month) <= 12:
        season = "Q4"
    print(stop_row)
    contract_number_col = None
    payment_date_col = None

    for col in range(1, sheet.max_column + 1):
        header_value = sheet.cell(row=1, column=col).value
        print(header_value)
        if header_value == "合同编号":
            contract_number_col = col
        elif header_value == "付款时间":
            payment_date_col = col

    # 确保找到了所需的列
    if contract_number_col is None or payment_date_col is None:
        print("未找到合同编号或付款日期列")
    else:
        for i in range(2, stop_row - minus):
            sheet.cell(row=i, column=int(contract_number_col)).value = contract_number
            sheet.cell(row=i, column=int(payment_date_col)).value = f"{year} {season}"
    # 填充数据
    workbook.save(f"{save_path.get().strip().replace("\"", "").replace("\"", "")}/output.xlsx")
    print("已导出")

workbook = ""
target_workbook = ""
sheet = ""
target_sheet = ""
cell = ""
target_cell = ""
index = []

root = tk.Tk()
module_path = tk.StringVar(value="C:/Users/25018/Documents/WeChat Files/wxid_w92198mya6n522/FileStorage/File/2024-10/模板.xlsx")
target_path = tk.StringVar(value="F:/shuju/2023年第三季度对账单/对账单（非框架）-邦汇保理专项_2023年第三季度_20230914(1).xlsx")
save_path = tk.StringVar(value="F:/0美术素材")
root.title("excel operator")
window_width = 300
window_height = 200
canvas = tk.Canvas(root, width=window_width, height=window_height)
canvas.grid(row=0, column=0, rowspan=50, columnspan=50)  # 覆盖整个窗口
tk.Label(root, text="在这里填写模板文件的路径").grid(row=1, column=0)
tk.Entry(root, textvariable=module_path).grid(row=1, column=1)
tk.Label(root, text="在这里填写目标文件的路径").grid(row=2, column=0)
tk.Entry(root, textvariable=target_path).grid(row=2, column=1)
tk.Label(root, text="保存路径").grid(row=3, column=0)
tk.Entry(root, textvariable=save_path).grid(row=3, column=1)
submit_btn = ttk.Button(root, text="执行格式转换", command=operate)
submit_btn.grid(row=9, column=0)
root.mainloop()
